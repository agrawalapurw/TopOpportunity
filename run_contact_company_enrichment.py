import argparse
from copy import copy
import fnmatch
import getpass
import json
import os
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional
from urllib.parse import urlparse

import pandas as pd
import pyodbc
import tldextract
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CONTACT_EMAIL_CANDIDATES = [
    "contact_email",
    "contact email",
    "email",
    "email_address",
    "email address",
]

WEBSITE_COLUMN_CANDIDATES = [
    "website",
    "company website",
    "web site",
    "url",
    "company_url",
    "domain",
    "website domain",
    "email_address_domain",
]

SHIP_TO_COMPANY_CANDIDATES = [
    "ship_to_company_name",
    "ship to company name",
    "ship_to_company",
    "ship to company",
    "company_name",
    "ship to customer name",
    "ship_to_name",
    "sold to customer name",
]

QUERY_TEMPLATE = """
select
A.email_address,
A.first_name,
A.last_name,
A.company,
A.city,
A.state_or_province,
A.zip_or_postal_code,
A.country,
A.business_phone,
A.mobile_phone,
A.title,
A.salutation,
A.salesperson,
A.lead_source_most_recent,
A.lead_source_original,
A.industry,
A.lead_status,
A.job_role,
A.region,
A.account_type,
A.contact_type,
A.country_code,
A.account_owner,
A.call_back_date,
A.call_back_time,
A.mkt_opt_in_request,
A.contact_status,
A.call_back_time,
A.mkt_opt_in_request,
A.contact_status,
A.contactid,
A.date_created,
A.date_modified,
A.eloqua_contact_id,
A.email_address_domain,
A.lead_trigger,
B.lead_triggering_activity,
B.row_num,
A.mkt_opt_in_date,
A.mkt_opt_out_date
from vdb_bl_sm_sales_marketing.bl_smsm_elq_base_contact A
left join
(SELECT eloqua_contact_id,
lead_triggering_activity,
ROW_NUMBER() OVER (PARTITION BY eloqua_contact_id ORDER BY TO_DATE(datacard_updated_at, 'YYYY-MM-DD') DESC) AS row_num
FROM vdb_bl_sm_sales_marketing.bl_smsm_elq_base_lead_lifecycle) as B
on A.eloqua_contact_id = B.eloqua_contact_id
where A.email_address_domain in ({placeholders})
and B.row_num = 1
and A.email_address is not null
order by A.email_address_domain asc
"""

# Output styling to mirror legacy enrichment visual format.
DOMAIN_HDR_FILL = PatternFill(fill_type="solid", fgColor="F4B942")
DOMAIN_EVEN_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
DOMAIN_ODD_FILL = PatternFill(fill_type="solid", fgColor="FFE8A1")
ENRICH_HDR_FILL = PatternFill(fill_type="solid", fgColor="2E75B6")
ENRICH_EVEN_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
ENRICH_ODD_FILL = PatternFill(fill_type="solid", fgColor="C5D3EC")
STATUS_MATCHED_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
STATUS_UNMATCHED_FILL = PatternFill(fill_type="solid", fgColor="FFCCCC")

HEADER_FONT_DARK = Font(bold=True, color="FFFFFF", size=10)
HEADER_FONT_AMBER = Font(bold=True, color="000000", size=10)
DATA_FONT = Font(size=10)

_THIN = Side(style="thin", color="B0B0B0")
CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTRE_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

ARG_DEFAULTS = {
    "input_dir": Path("input"),
    "glob": "*.xlsx",
    "driver": "auto",
    "server": "dive-dev.infineon.com",
    "port": 9996,
    "database": "vdb_bl_sm_web",
    "uid": os.getenv("DIVE_UID", ""),
    "password": os.getenv("DIVE_PWD", ""),
    "chunk_size": 1000,
    "output_dir": Path("output"),
    "workbooks": None,
}


@dataclass
class CompanyMatch:
    company: Optional[str]
    method: str
    score: Optional[float]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read website/domain/contact_email from Excel, query DIVE by domain, "
            "and append best matched company name."
        )
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config.json"),
        help="Optional JSON config file path. CLI args override config values.",
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=ARG_DEFAULTS["input_dir"],
        help="Directory containing source Excel files.",
    )
    parser.add_argument(
        "--glob",
        default=ARG_DEFAULTS["glob"],
        help="Glob pattern for input files.",
    )
    parser.add_argument(
        "--driver",
        default=ARG_DEFAULTS["driver"],
        help="ODBC driver name. Use 'auto' to detect an installed Denodo driver.",
    )
    parser.add_argument("--server", default=ARG_DEFAULTS["server"], help="DIVE server hostname.")
    parser.add_argument("--port", type=int, default=ARG_DEFAULTS["port"], help="DIVE ODBC port.")
    parser.add_argument("--database", default=ARG_DEFAULTS["database"], help="DIVE database.")
    parser.add_argument(
        "--uid",
        default=ARG_DEFAULTS["uid"],
        help="DIVE UID. If omitted, uses DIVE_UID environment variable.",
    )
    parser.add_argument(
        "--password",
        default=ARG_DEFAULTS["password"],
        help="DIVE password. If omitted, uses DIVE_PWD environment variable or secure prompt.",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=ARG_DEFAULTS["chunk_size"],
        help="Number of domains per SQL IN chunk.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=ARG_DEFAULTS["output_dir"],
        help="Directory for enriched files.",
    )
    parser.add_argument(
        "--workbook",
        dest="workbooks",
        action="append",
        default=ARG_DEFAULTS["workbooks"],
        help=(
            "Workbook file name or glob pattern to process. "
            "Repeat flag for multiple selections. Example: --workbook '*Input*.xlsx'"
        ),
    )
    return parser.parse_args()


def load_config(config_path: Path) -> dict:
    if not config_path.exists():
        return {}
    with config_path.open("r", encoding="utf-8") as fp:
        return json.load(fp)


def apply_config_defaults(args: argparse.Namespace, config: dict) -> argparse.Namespace:
    path_keys = {"input_dir", "output_dir"}
    for key, default_value in ARG_DEFAULTS.items():
        if key not in config:
            continue
        current_value = getattr(args, key)
        if current_value != default_value:
            continue
        value = config[key]
        if key in path_keys and isinstance(value, str):
            value = Path(value)
        if key == "workbooks" and isinstance(value, str):
            value = [value]
        setattr(args, key, value)
    return args


def resolve_odbc_driver(requested_driver: str) -> str:
    available = pyodbc.drivers()
    if not available:
        raise RuntimeError("No ODBC drivers found on this system. Install the Denodo ODBC driver first.")

    requested = (requested_driver or "").strip()
    if requested and requested.lower() != "auto":
        if requested in available:
            return requested
        matching = [d for d in available if requested.lower() in d.lower()]
        if matching:
            print(f"Requested driver '{requested}' not found exactly. Using '{matching[0]}'")
            return matching[0]
        raise RuntimeError(f"Requested ODBC driver '{requested}' is not installed. Available drivers: {available}")

    denodo_drivers = [d for d in available if "denodo" in d.lower()]
    if denodo_drivers:
        preferred_order = sorted(
            denodo_drivers,
            key=lambda d: ("x64" not in d.lower(), "unicode" not in d.lower(), d.lower()),
        )
        print(f"Auto-selected ODBC driver: {preferred_order[0]}")
        return preferred_order[0]

    raise RuntimeError(f"No Denodo ODBC driver found. Available drivers now: {available}")


def build_connection(args: argparse.Namespace) -> pyodbc.Connection:
    selected_driver = resolve_odbc_driver(args.driver)
    uid = args.uid or input("DIVE UID: ").strip()
    pwd = args.password or getpass.getpass("DIVE password: ")

    conn_string = (
        f"DRIVER={{{selected_driver}}};"
        f"DATABASE={args.database};"
        f"SERVER={args.server};"
        f"PORT={args.port};"
        f"UID={uid};"
        f"PWD={pwd}"
    )
    return pyodbc.connect(conn_string)


def find_matching_column(df: pd.DataFrame, candidates: list[str], label: str) -> str:
    normalized = {_normalize_header(c): c for c in df.columns}
    candidate_keys = [_normalize_header(c) for c in candidates]

    for candidate in candidate_keys:
        if candidate in normalized:
            return normalized[candidate]

    # Fallback: substring match for noisy headers like "Contact Email Address (Primary)"
    for candidate in candidate_keys:
        for norm_key, original in normalized.items():
            if candidate and candidate in norm_key:
                return original

    raise ValueError(f"Could not find required column for {label}. Present columns: {list(df.columns)}")


def try_find_matching_column(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    try:
        return find_matching_column(df, candidates, "optional field")
    except ValueError:
        return None


def _normalize_header(value: object) -> str:
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def detect_header_row(workbook_path: Path, sheet_name: str, scan_rows: int = 40) -> int:
    raw = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, nrows=scan_rows)

    possible_email = {_normalize_header(c) for c in CONTACT_EMAIL_CANDIDATES}
    possible_website = {_normalize_header(c) for c in WEBSITE_COLUMN_CANDIDATES}
    optional_ship = {_normalize_header(c) for c in SHIP_TO_COMPANY_CANDIDATES}

    best_idx = 0
    best_score = -1

    for idx, row in raw.iterrows():
        normalized_cells = {
            _normalize_header(v)
            for v in row.tolist()
            if pd.notna(v) and str(v).strip()
        }

        # Score rows that likely represent headers; email match is most important.
        score = 0
        if normalized_cells & possible_email:
            score += 3
        if normalized_cells & possible_website:
            score += 3
        if normalized_cells & optional_ship:
            score += 2
        score += int(pd.Series(row).notna().sum() > 2)

        if score > best_score:
            best_score = score
            best_idx = int(idx)

    return best_idx


def read_sheet_with_detected_header(workbook_path: Path, sheet_name: str) -> pd.DataFrame:
    header_row = detect_header_row(workbook_path, sheet_name)
    return pd.read_excel(workbook_path, sheet_name=sheet_name, header=header_row)


def normalize_email(email_value: object) -> Optional[str]:
    if pd.isna(email_value):
        return None
    raw = str(email_value).strip().lower()
    if not raw or "@" not in raw:
        return None
    return raw


def extract_domain_from_email(email_value: object) -> Optional[str]:
    normalized_email = normalize_email(email_value)
    if not normalized_email:
        return None
    parts = normalized_email.split("@")
    if len(parts) != 2:
        return None
    domain = parts[1].strip().strip(".")
    return domain or None


def canonicalize_domain(url_or_domain_value: object) -> Optional[str]:
    if pd.isna(url_or_domain_value):
        return None

    raw = str(url_or_domain_value).strip().lower()
    if not raw or raw in {"na", "n/a", "none", "null"}:
        return None

    parse_target = raw if "://" in raw else f"http://{raw}"
    parsed = urlparse(parse_target)
    host = parsed.netloc or parsed.path
    host = host.split("@")[ -1].split(":")[0].strip(".")
    if not host:
        return None

    extracted = tldextract.extract(host)
    if extracted.domain and extracted.suffix:
        return f"{extracted.domain}.{extracted.suffix}".lower()

    return host.lower()


def normalize_company_name(name_value: object) -> str:
    if pd.isna(name_value):
        return ""
    text = str(name_value).strip().lower()
    for ch in [",", ".", "-", "_", "/", "\\", "&", "(", ")"]:
        text = text.replace(ch, " ")
    return " ".join(text.split())


def chunked(items: list[str], size: int) -> Iterable[list[str]]:
    for i in range(0, len(items), size):
        yield items[i : i + size]


def timestamped_path(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


def _clean_cell_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _status_fill(value: object) -> PatternFill:
    v = str(value).strip().lower() if value is not None else ""
    if v == "company":
        return STATUS_MATCHED_FILL
    return STATUS_UNMATCHED_FILL


def _clone_workbook_cell_formatting_only(source_path: Path, target_path: Path) -> None:
    src_wb = load_workbook(source_path)
    new_wb = load_workbook(source_path)

    # Rebuild sheets with values/styles to avoid fragile table/drawing metadata issues.
    for sheet_name in list(new_wb.sheetnames):
        del new_wb[sheet_name]

    for src_ws in src_wb.worksheets:
        dst_ws = new_wb.create_sheet(src_ws.title)

        dst_ws.sheet_format = copy(src_ws.sheet_format)
        dst_ws.sheet_properties = copy(src_ws.sheet_properties)
        dst_ws.page_margins = copy(src_ws.page_margins)
        dst_ws.page_setup = copy(src_ws.page_setup)
        dst_ws.print_options = copy(src_ws.print_options)
        dst_ws.freeze_panes = src_ws.freeze_panes

        for key, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[key].width = dim.width
            dst_ws.column_dimensions[key].hidden = dim.hidden
            dst_ws.column_dimensions[key].outlineLevel = dim.outlineLevel

        for idx, dim in src_ws.row_dimensions.items():
            dst_ws.row_dimensions[idx].height = dim.height
            dst_ws.row_dimensions[idx].hidden = dim.hidden
            dst_ws.row_dimensions[idx].outlineLevel = dim.outlineLevel

        for row in src_ws.iter_rows(
            min_row=1,
            max_row=src_ws.max_row,
            min_col=1,
            max_col=src_ws.max_column,
        ):
            for cell in row:
                new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell._style = copy(cell._style)
                new_cell.number_format = cell.number_format
                new_cell.alignment = copy(cell.alignment)
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.protection = copy(cell.protection)

        for merged in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged))

    new_wb.save(target_path)


def write_enrichment_into_formatted_workbook(
    workbook_path: Path,
    enriched_sheets: dict[str, pd.DataFrame],
    all_matches_df: pd.DataFrame,
    output_workbook: Path,
) -> None:
    _clone_workbook_cell_formatting_only(workbook_path, output_workbook)
    wb = load_workbook(output_workbook)

    for sheet_name, enriched_df in enriched_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        header_row = detect_header_row(workbook_path, sheet_name) + 1

        existing_headers: list[str] = []
        max_existing_col = ws.max_column
        for col in range(1, max_existing_col + 1):
            value = ws.cell(row=header_row, column=col).value
            existing_headers.append(str(value).strip() if value is not None else "")

        preferred_order = [
            "input_domain",
            "matched_company_name",
            "company_match_method",
            "company_match_score",
            "is_company",
        ]
        new_columns = [c for c in preferred_order if c in enriched_df.columns and c not in existing_headers]
        if not new_columns:
            continue

        num_data_rows = len(enriched_df)
        start_col = max_existing_col + 1

        for offset, col_name in enumerate(new_columns):
            col_idx = start_col + offset
            col_letter = get_column_letter(col_idx)
            is_domain = col_name == "input_domain"
            is_status = col_name == "is_company"

            header_cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            header_cell.fill = DOMAIN_HDR_FILL if is_domain else ENRICH_HDR_FILL
            header_cell.font = HEADER_FONT_AMBER if is_domain else HEADER_FONT_DARK
            header_cell.border = CELL_BORDER
            header_cell.alignment = CENTRE_ALIGN

            col_values = enriched_df[col_name].tolist()
            max_len = len(col_name)
            for row_offset, value in enumerate(col_values, start=1):
                cleaned = _clean_cell_value(value)
                cell = ws.cell(row=header_row + row_offset, column=col_idx, value=cleaned)
                cell.font = DATA_FONT
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                if is_status:
                    cell.fill = _status_fill(cleaned)
                elif is_domain:
                    cell.fill = DOMAIN_EVEN_FILL if row_offset % 2 == 0 else DOMAIN_ODD_FILL
                else:
                    cell.fill = ENRICH_EVEN_FILL if row_offset % 2 == 0 else ENRICH_ODD_FILL

                if cleaned is not None:
                    max_len = max(max_len, len(str(cleaned)))

            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 48)

        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_existing_col + len(new_columns))}{header_row}"

    # Add dedicated detailed match tab with one row per input-to-DIVE domain match pair.
    all_matches_sheet_name = "Data_All Matches"
    if all_matches_sheet_name in wb.sheetnames:
        del wb[all_matches_sheet_name]

    ws_all = wb.create_sheet(all_matches_sheet_name)
    if all_matches_df.empty:
        headers = ["sheet_name", "input_domain", "matched_company_name", "company_match_method"]
        for col_idx, header in enumerate(headers, start=1):
            h = ws_all.cell(row=1, column=col_idx, value=header)
            h.font = HEADER_FONT_DARK
            h.fill = ENRICH_HDR_FILL
            h.border = CELL_BORDER
            h.alignment = CENTRE_ALIGN
            ws_all.column_dimensions[get_column_letter(col_idx)].width = max(18, len(header) + 3)
    else:
        headers = [str(c) for c in all_matches_df.columns]
        for col_idx, header in enumerate(headers, start=1):
            h = ws_all.cell(row=1, column=col_idx, value=header)
            h.font = HEADER_FONT_DARK
            h.fill = ENRICH_HDR_FILL
            h.border = CELL_BORDER
            h.alignment = CENTRE_ALIGN

        for row_idx, row_values in enumerate(all_matches_df.itertuples(index=False, name=None), start=2):
            for col_idx, raw_value in enumerate(row_values, start=1):
                cleaned = _clean_cell_value(raw_value)
                cell = ws_all.cell(row=row_idx, column=col_idx, value=cleaned)
                cell.font = DATA_FONT
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.fill = ENRICH_EVEN_FILL if row_idx % 2 == 0 else ENRICH_ODD_FILL

        for col_idx, header in enumerate(headers, start=1):
            series = all_matches_df.iloc[:, col_idx - 1].dropna().astype(str)
            max_len = max([len(str(header))] + ([int(series.str.len().max())] if not series.empty else [0]))
            ws_all.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 14), 48)

    ws_all.freeze_panes = "A2"
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output_workbook)


def fetch_dive_matches(conn: pyodbc.Connection, domains: list[str], chunk_size: int) -> pd.DataFrame:
    if not domains:
        return pd.DataFrame(columns=["email_address_domain", "email_address", "company"])

    all_frames: list[pd.DataFrame] = []
    for domain_chunk in chunked(domains, chunk_size):
        placeholders = ",".join("?" for _ in domain_chunk)
        sql = QUERY_TEMPLATE.format(placeholders=placeholders)
        frame = pd.read_sql_query(sql, conn, params=domain_chunk)
        all_frames.append(frame)

    if not all_frames:
        return pd.DataFrame(columns=["email_address_domain", "email_address", "company"])

    combined = pd.concat(all_frames, ignore_index=True)
    combined["email_address_domain"] = combined["email_address_domain"].astype(str).str.lower()
    combined["email_address"] = combined["email_address"].astype(str).str.lower()
    combined["company"] = combined["company"].astype(str)
    return combined


def build_lookup_tables(matches: pd.DataFrame) -> tuple[dict[tuple[str, str], str], dict[str, list[tuple[str, int]]]]:
    email_company: dict[tuple[str, str], str] = {}
    domain_company: dict[str, list[tuple[str, int]]] = {}

    if matches.empty:
        return email_company, domain_company

    safe = matches.dropna(subset=["email_address_domain", "company"]).copy()
    safe["email_key"] = safe["email_address"].fillna("").astype(str).str.strip().str.lower()

    email_counts = (
        safe[safe["email_key"] != ""]
        .groupby(["email_address_domain", "email_key", "company"])
        .size()
        .reset_index(name="cnt")
    )
    for (domain, email_key), grp in email_counts.groupby(["email_address_domain", "email_key"]):
        top = grp.sort_values(["cnt", "company"], ascending=[False, True]).iloc[0]
        email_company[(str(domain), str(email_key))] = str(top["company"])

    domain_counts = (
        safe.groupby(["email_address_domain", "company"]).size().reset_index(name="cnt")
    )
    for domain, grp in domain_counts.groupby("email_address_domain"):
        ordered = (
            grp.sort_values(["cnt", "company"], ascending=[False, True])
            [["company", "cnt"]]
            .itertuples(index=False, name=None)
        )
        domain_company[str(domain)] = [(str(company), int(cnt)) for company, cnt in ordered]

    return email_company, domain_company


def choose_company(
    email: Optional[str],
    domain: Optional[str],
    ship_to_company: object,
    email_company_lookup: dict[tuple[str, str], str],
    domain_company_lookup: dict[str, list[tuple[str, int]]],
) -> CompanyMatch:
    if not domain:
        return CompanyMatch(company=None, method="no_input_domain", score=None)

    if email:
        exact_key = (domain, email)
        if exact_key in email_company_lookup:
            return CompanyMatch(company=email_company_lookup[exact_key], method="exact_email", score=1.0)

    candidates = domain_company_lookup.get(domain, [])
    if not candidates:
        return CompanyMatch(company=None, method="no_domain_match", score=None)

    normalized_ship = normalize_company_name(ship_to_company)
    if normalized_ship:
        best_company = None
        best_score = -1.0
        for company, _ in candidates:
            candidate_norm = normalize_company_name(company)
            score = SequenceMatcher(None, normalized_ship, candidate_norm).ratio()
            if score > best_score:
                best_score = score
                best_company = company
        if best_company is not None:
            return CompanyMatch(
                company=best_company,
                method="domain_plus_ship_to",
                score=round(float(best_score), 4),
            )

    return CompanyMatch(company=candidates[0][0], method="domain_frequency", score=None)


def filter_workbooks(workbooks: list[Path], selected_patterns: Optional[list[str]]) -> list[Path]:
    if not selected_patterns:
        return workbooks

    selected: list[Path] = []
    seen: set[Path] = set()
    for workbook in workbooks:
        for pattern in selected_patterns:
            if fnmatch.fnmatch(workbook.name, pattern):
                if workbook not in seen:
                    selected.append(workbook)
                    seen.add(workbook)
                break
    return selected


def pick_workbooks_in_terminal(all_workbooks: list[Path]) -> list[Path]:
    if not all_workbooks:
        return []

    if len(all_workbooks) == 1:
        return all_workbooks

    print("\nSelect workbook(s) to process:")
    for idx, wb in enumerate(all_workbooks, start=1):
        print(f"  {idx}. {wb.name}")
    print("Enter numbers separated by commas (example: 1,3,5) or 'all'.")

    try:
        choice = input("Selection [all]: ").strip().lower()
    except EOFError:
        print(
            "Interactive selection is unavailable in this run context (stdin EOF). "
            "Re-run in a terminal or pass one/more --workbook arguments."
        )
        return []

    if not choice or choice == "all":
        return all_workbooks

    selected: list[Path] = []
    for token in choice.split(","):
        token = token.strip()
        if not token:
            continue
        if token.isdigit():
            i = int(token)
            if 1 <= i <= len(all_workbooks):
                selected.append(all_workbooks[i - 1])

    deduped: list[Path] = []
    seen: set[Path] = set()
    for wb in selected:
        if wb not in seen:
            deduped.append(wb)
            seen.add(wb)

    if not deduped:
        print("No valid selection entered. No files will be processed.")
    return deduped


def process_sheet(df: pd.DataFrame, matches: pd.DataFrame) -> pd.DataFrame:
    email_col = try_find_matching_column(df, CONTACT_EMAIL_CANDIDATES)
    website_col = try_find_matching_column(df, WEBSITE_COLUMN_CANDIDATES)
    ship_col = try_find_matching_column(df, SHIP_TO_COMPANY_CANDIDATES)

    output = df.copy()
    if email_col is not None:
        output["_input_email"] = output[email_col].apply(normalize_email)
        output["input_domain"] = output[email_col].apply(extract_domain_from_email)
    else:
        output["_input_email"] = None
        output["input_domain"] = None

    if website_col is not None:
        website_domains = output[website_col].apply(canonicalize_domain)
        output["input_domain"] = output["input_domain"].fillna(website_domains)

    if ship_col is None:
        output["_ship_to_company_name_input"] = ""
        ship_values = output["_ship_to_company_name_input"]
    else:
        ship_values = output[ship_col]

    email_lookup, domain_lookup = build_lookup_tables(matches)

    results = [
        choose_company(email, domain, ship_name, email_lookup, domain_lookup)
        for email, domain, ship_name in zip(output["_input_email"], output["input_domain"], ship_values)
    ]

    output["matched_company_name"] = [r.company for r in results]
    output["company_match_method"] = [r.method for r in results]
    output["company_match_score"] = [r.score for r in results]
    output["is_company"] = output["matched_company_name"].apply(
        lambda v: "company" if pd.notna(v) and str(v).strip() else "not_company"
    )

    drop_cols = ["_input_email"]
    if "_ship_to_company_name_input" in output.columns:
        drop_cols.append("_ship_to_company_name_input")
    output = output.drop(columns=drop_cols)
    return output


def build_all_matches_sheet(
    sheet_name: str,
    enriched_df: pd.DataFrame,
    matches: pd.DataFrame,
) -> pd.DataFrame:
    base = enriched_df.copy()
    base.insert(0, "sheet_name", sheet_name)
    base["_row_id"] = range(1, len(base) + 1)

    merged = base.merge(
        matches,
        how="left",
        left_on="input_domain",
        right_on="email_address_domain",
        suffixes=("", "_dive"),
    )

    preferred_columns = [
        "sheet_name",
        "_row_id",
        "input_domain",
        "matched_company_name",
        "company_match_method",
        "company_match_score",
        "is_company",
        "email_address",
        "company",
        "email_address_domain",
    ]
    ordered = [c for c in preferred_columns if c in merged.columns]
    remaining = [c for c in merged.columns if c not in ordered]
    merged = merged[ordered + remaining]
    merged = merged.rename(columns={"_row_id": "source_row_number", "company": "dive_company"})
    return merged


def process_workbook(workbook_path: Path, conn: pyodbc.Connection, output_dir: Path, chunk_size: int) -> dict[str, int]:
    xls = pd.ExcelFile(workbook_path)
    sheets: dict[str, pd.DataFrame] = {}
    domains: set[str] = set()

    for sheet_name in xls.sheet_names:
        df = read_sheet_with_detected_header(workbook_path, sheet_name)

        email_col = try_find_matching_column(df, CONTACT_EMAIL_CANDIDATES)
        website_col = try_find_matching_column(df, WEBSITE_COLUMN_CANDIDATES)
        if email_col is None and website_col is None:
            continue

        sheet_domains: list[str] = []
        if email_col is not None:
            sheet_domains.extend(
                df[email_col].apply(extract_domain_from_email).dropna().astype(str).str.lower().tolist()
            )
        if website_col is not None:
            sheet_domains.extend(
                df[website_col].apply(canonicalize_domain).dropna().astype(str).str.lower().tolist()
            )
        sheet_domains = sorted(set(sheet_domains))
        if sheet_domains:
            domains.update(sheet_domains)
        sheets[sheet_name] = df

    if not sheets:
        print(
            f"Skipping {workbook_path.name}: no sheet contains recognizable website/domain/contact_email columns."
        )
        return {
            "sheets_processed": 0,
            "unique_domains": 0,
            "dive_rows_returned": 0,
        }

    domain_list = sorted(domains)
    matches = fetch_dive_matches(conn, domain_list, chunk_size)

    enriched_sheets: dict[str, pd.DataFrame] = {}
    all_matches_frames: list[pd.DataFrame] = []
    unmatched_rows: list[pd.DataFrame] = []
    for sheet_name, df in sheets.items():
        enriched = process_sheet(df, matches)
        enriched_sheets[sheet_name] = enriched
        all_matches_frames.append(build_all_matches_sheet(sheet_name, enriched, matches))
        unmatched = enriched[enriched["matched_company_name"].isna()].copy()
        if not unmatched.empty:
            unmatched.insert(0, "sheet_name", sheet_name)
            unmatched_rows.append(unmatched)

    all_matches_df = pd.concat(all_matches_frames, ignore_index=True) if all_matches_frames else pd.DataFrame()

    output_dir.mkdir(parents=True, exist_ok=True)
    output_workbook = output_dir / f"{workbook_path.stem}_company_enriched.xlsx"
    try:
        write_enrichment_into_formatted_workbook(
            workbook_path=workbook_path,
            enriched_sheets=enriched_sheets,
            all_matches_df=all_matches_df,
            output_workbook=output_workbook,
        )
    except PermissionError:
        fallback_output_workbook = timestamped_path(output_workbook)
        print(
            f"Output workbook is locked: {output_workbook.name}. "
            f"Writing to fallback file: {fallback_output_workbook.name}"
        )
        write_enrichment_into_formatted_workbook(
            workbook_path=workbook_path,
            enriched_sheets=enriched_sheets,
            all_matches_df=all_matches_df,
            output_workbook=fallback_output_workbook,
        )

    unmatched_path = output_dir / f"{workbook_path.stem}_unmatched_contacts.csv"
    try:
        if unmatched_rows:
            pd.concat(unmatched_rows, ignore_index=True).to_csv(unmatched_path, index=False)
        else:
            pd.DataFrame(columns=["sheet_name", "input_domain", "matched_company_name"]).to_csv(unmatched_path, index=False)
    except PermissionError:
        fallback_unmatched_path = timestamped_path(unmatched_path)
        print(
            f"Unmatched report is locked: {unmatched_path.name}. "
            f"Writing to fallback file: {fallback_unmatched_path.name}"
        )
        if unmatched_rows:
            pd.concat(unmatched_rows, ignore_index=True).to_csv(fallback_unmatched_path, index=False)
        else:
            pd.DataFrame(columns=["sheet_name", "input_domain", "matched_company_name"]).to_csv(fallback_unmatched_path, index=False)

    return {
        "sheets_processed": len(sheets),
        "unique_domains": len(domain_list),
        "dive_rows_returned": int(len(matches)),
    }


def main() -> None:
    args = parse_args()
    config_path = args.config
    if not config_path.is_absolute():
        config_path = (Path(__file__).parent / config_path).resolve()
    args = apply_config_defaults(args, load_config(config_path))

    input_dir = args.input_dir
    if not input_dir.is_absolute():
        input_dir = (Path(__file__).parent / input_dir).resolve()

    output_dir = args.output_dir
    if not output_dir.is_absolute():
        output_dir = (Path(__file__).parent / output_dir).resolve()

    if not input_dir.exists():
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    all_workbooks = [
        p
        for p in sorted(input_dir.glob(args.glob))
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"} and not p.name.startswith("~$")
    ]
    if args.workbooks:
        workbooks = filter_workbooks(all_workbooks, args.workbooks)
    else:
        workbooks = pick_workbooks_in_terminal(all_workbooks)

    if not workbooks:
        if args.workbooks:
            print(f"No matching workbooks found in {input_dir} for patterns: {', '.join(args.workbooks)}")
        else:
            print(f"No Excel files found in {input_dir} with pattern {args.glob}")
        return

    print(f"Found {len(workbooks)} workbook(s) in {input_dir}")

    conn = build_connection(args)
    try:
        run_summary: list[dict[str, object]] = []
        for workbook in workbooks:
            print(f"Processing: {workbook.name}")
            stats = process_workbook(
                workbook_path=workbook,
                conn=conn,
                output_dir=output_dir,
                chunk_size=args.chunk_size,
            )
            run_summary.append(
                {
                    "workbook": workbook.name,
                    **stats,
                }
            )

        summary_file = output_dir / "run_summary.csv"
        final_summary_file = summary_file
        try:
            pd.DataFrame(run_summary).to_csv(summary_file, index=False)
        except PermissionError:
            fallback_summary_file = timestamped_path(summary_file)
            print(
                f"Summary file is locked: {summary_file.name}. "
                f"Writing to fallback file: {fallback_summary_file.name}"
            )
            pd.DataFrame(run_summary).to_csv(fallback_summary_file, index=False)
            final_summary_file = fallback_summary_file
        print(f"Completed. Outputs written to: {output_dir}")
        print(f"Summary: {final_summary_file.name}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
